import openpyxl
from db import SQLiteDB
import os


def write_excel_file(name: str, table_data: list):
    file_name = f"output/{name}.xlsx"
    if os.path.exists(file_name):
        # 既存のファイルを読み込む
        wb = openpyxl.load_workbook(file_name)
    else:
        # 新規作成する
        wb = openpyxl.Workbook()

    # アクティブなワークシートを選択する
    sheet = wb.active

    sheet.cell(row=1, column=1, value="ask")
    sheet.cell(row=1, column=2, value="data")

    # 行番号をカウントする変数を定義する
    row_number = 2  # 2行目から書き込む（1行目はヘッダー）

    # シンボル情報リストの各要素を書き込む
    for symbol in table_data:
        ask = symbol[0]
        data = symbol[1]

        # 行に情報を書き込む
        sheet.cell(row=row_number, column=1, value=ask)
        sheet.cell(row=row_number, column=2, value=data)

        # 行番号を増やす
        row_number += 1

    # Excelファイルを保存する
    wb.save(f"output/{name}.xlsx")


def db_to_excel_file(name):
    # dbから特定のテーブルを参照して、Excelファイルに出力する
    # SQLiteDBインスタンスの生成
    db = SQLiteDB("db/symbols_info.db")
    # 特定のテーブルのデータを取得
    table_data = db.fetch_data(f"SELECT * FROM {name}")
    # Excelファイルに出力
    write_excel_file(name, table_data)
    # データベースを閉じる
    db.close()


if __name__ == "__main__":
    pass
